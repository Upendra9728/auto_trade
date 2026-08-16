"""Shared helper for building downloadable .xlsx exports (admin reports)."""
from __future__ import annotations

import datetime as dt
import io
from typing import Any

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_IST_OFFSET = dt.timedelta(hours=5, minutes=30)


# ── Colour palette ────────────────────────────────────────────────────────────
# Only two colours are used for data rows:
#   - White
#   - Light grey
#
# Header has its own colour and is not part of row grouping.
_HDR_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_ROW_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFF")
_ROW_GREY = PatternFill(fill_type="solid", fgColor="F2F2F2")


def to_ist_str(value: dt.datetime | None) -> str:
    """Formats a UTC-naive datetime as an IST 'YYYY-MM-DD HH:MM:SS' string."""
    if value is None:
        return ""

    return (value + _IST_OFFSET).strftime("%Y-%m-%d %H:%M:%S")


def build_signal_group_fills(
    signal_ids: list[Any],
) -> list[PatternFill]:
    """
    Creates row fills based on Signal ID groups.

    Example:
        Signal IDs:
            101
            101
            101
            102
            102
            103
            103

        Colours:
            White
            White
            White
            Grey
            Grey
            White
            White

    The colour changes only when the Signal ID changes.
    """

    fills: list[PatternFill] = []

    if not signal_ids:
        return fills

    previous_signal_id = signal_ids[0]
    use_grey = False

    for signal_id in signal_ids:
        # Change colour only when Signal ID changes.
        if signal_id != previous_signal_id:
            use_grey = not use_grey
            previous_signal_id = signal_id

        fills.append(
            _ROW_GREY if use_grey else _ROW_WHITE
        )

    return fills


def build_xlsx_response(
    *,
    filename: str,
    sheet_title: str,
    headers: list[str],
    rows: list[list],
    signal_id_column: int | None = None,
) -> StreamingResponse:
    """
    Builds a single-sheet .xlsx and returns it as a download.

    Args:
        filename:
            Name of the downloaded Excel file.

        sheet_title:
            Excel sheet name.

        headers:
            List of column headers.

        rows:
            Excel data rows.

        signal_id_column:
            Zero-based index of the Signal ID column.

            Example:
                signal_id_column=0
                -> Signal ID is the first column.

                signal_id_column=3
                -> Signal ID is the fourth column.

            If None, rows will simply use alternating white/grey colours.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    # ── Header ────────────────────────────────────────────────────────────────
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.row_dimensions[1].height = 30

    # ── Prepare Signal ID based fills ─────────────────────────────────────────
    row_fills: list[PatternFill]

    if signal_id_column is not None:
        signal_ids = [
            row[signal_id_column]
            if signal_id_column < len(row)
            else None
            for row in rows
        ]

        row_fills = build_signal_group_fills(signal_ids)

    else:
        # Fallback: alternate white/grey per row.
        row_fills = [
            _ROW_WHITE if index % 2 == 0 else _ROW_GREY
            for index in range(len(rows))
        ]

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_index, row in enumerate(rows, start=2):
        ws.append(row)

        fill = row_fills[row_index - 2]

        for cell in ws[row_index]:
            cell.fill = fill

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_index, header in enumerate(headers, start=1):
        width = max(
            12,
            min(
                42,
                len(str(header)) + 4,
            ),
        )

        column_letter = ws.cell(
            row=1,
            column=col_index,
        ).column_letter

        ws.column_dimensions[column_letter].width = width

    # ── Create Excel file in memory ───────────────────────────────────────────
    buffer = io.BytesIO()

    wb.save(buffer)
    buffer.seek(0)

    # ── Return downloadable response ─────────────────────────────────────────
    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        },
    )